from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_EXCEL_PATH = Path(r"C:\Users\Youss\OneDrive\Bureau\Fiche de Aromair.xlsx")
DEFAULT_TARGET_DIRS = (
    Path("data/processed/maintenance"),
    Path("data/processed/Maintenance_json"),
)


@dataclass(frozen=True)
class ReferenceColumn:
    header: str
    key: str


REFERENCE_COLUMNS = (
    ReferenceColumn("Liste des clients", "clients"),
    ReferenceColumn("Liste des parfums", "parfums"),
    ReferenceColumn("Liste des adresses", "addresses"),
    ReferenceColumn("Liste des adresses ", "addresses"),
    ReferenceColumn("Liste des modeles diffuseurs", "models"),
    ReferenceColumn("Liste des Ref", "diffuser_refs"),
    ReferenceColumn("Liste des noms des techniciens", "technicians"),
    ReferenceColumn("Liste des noms des techniciens ", "technicians"),
    ReferenceColumn("Liste des emplacements", "emplacements"),
    ReferenceColumn("Liste des references bouteilles", "bottle_refs"),
)


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    lowered = ascii_text.lower()
    collapsed = "".join(character for character in lowered if character.isalnum())
    return collapsed


def similarity(left: str, right: str) -> float:
    return SequenceMatcher(a=normalize_text(left), b=normalize_text(right)).ratio()


def variants(value: str | None) -> list[str]:
    if value is None:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    output = [raw]
    output.extend(match.strip() for match in re.findall(r"\(([^)]+)\)", raw) if match.strip())
    output.extend(match.strip() for match in re.findall(r"\b[A-Z]{1,5}\b", raw) if match.strip())
    unique: list[str] = []
    seen: set[str] = set()
    for item in output:
        key = normalize_text(item)
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(item)
    return unique


def load_excel_references(excel_path: Path) -> dict[str, list[str]]:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    column_lookup = {column.header: column.key for column in REFERENCE_COLUMNS}
    references: dict[str, list[str]] = {}

    for column_index, header in enumerate(header_row):
        key = column_lookup.get(header)
        if not key:
            continue
        values: list[str] = []
        for row in rows[1:]:
            if column_index >= len(row):
                continue
            cell = row[column_index]
            if cell is None:
                continue
            value = str(cell).strip()
            if not value:
                continue
            if value not in values:
                values.append(value)
        references[key] = values
    return references


def choose_reference(value: str | None, candidates: list[str], threshold: float = 0.84) -> tuple[str | None, float]:
    if value is None:
        return None, 0.0
    raw = str(value).strip()
    if not raw:
        return None, 0.0

    input_variants = variants(raw)
    normalized_input_variants = {normalize_text(item) for item in input_variants}
    for candidate in candidates:
        candidate_variants = variants(candidate)
        if any(normalize_text(item) in normalized_input_variants for item in candidate_variants):
            return candidate, 1.0

    best_candidate: str | None = None
    best_score = 0.0
    for candidate in candidates:
        candidate_variants = variants(candidate)
        score = max(
            similarity(left_variant, right_variant)
            for left_variant in input_variants
            for right_variant in candidate_variants
        )
        if score > best_score:
            best_candidate = candidate
            best_score = score
    if best_candidate is None or best_score < threshold:
        return None, best_score
    return best_candidate, best_score


def get_from_path(payload: dict[str, Any], path: list[str | int]) -> Any:
    current: Any = payload
    for key in path:
        if isinstance(key, int):
            if not isinstance(current, list) or key >= len(current):
                return None
            current = current[key]
        else:
            if not isinstance(current, dict):
                return None
            current = current.get(key)
    return current


def set_in_path(payload: dict[str, Any], path: list[str | int], value: Any) -> bool:
    current: Any = payload
    for key in path[:-1]:
        if isinstance(key, int):
            if not isinstance(current, list) or key >= len(current):
                return False
            current = current[key]
        else:
            if not isinstance(current, dict) or key not in current:
                return False
            current = current[key]
    last_key = path[-1]
    if isinstance(last_key, int):
        if not isinstance(current, list) or last_key >= len(current):
            return False
        current[last_key] = value
        return True
    if not isinstance(current, dict):
        return False
    current[last_key] = value
    return True


def iter_candidate_paths(payload: dict[str, Any]) -> list[tuple[str, list[str | int]]]:
    paths: list[tuple[str, list[str | int]]] = []
    paths.append(("clients", ["maintenance_details", "client"]))
    paths.append(("addresses", ["maintenance_details", "address"]))
    paths.append(("technicians", ["maintenance_details", "technician_name"]))
    paths.append(("clients", ["raw_payload", "maintenance_details", "client"]))
    paths.append(("addresses", ["raw_payload", "maintenance_details", "address"]))
    paths.append(("technicians", ["raw_payload", "maintenance_details", "technician_name"]))

    controls = payload.get("controle_diffuseur_recharge")
    if isinstance(controls, list):
        for index, _item in enumerate(controls):
            base = ["controle_diffuseur_recharge", index]
            raw_base = ["raw_payload", "controle_diffuseur_recharge", index]
            paths.extend(
                [
                    ("models", [*base, "model_diffuseur"]),
                    ("diffuser_refs", [*base, "reference_diffuseur"]),
                    ("parfums", [*base, "nom_parfum"]),
                    ("emplacements", [*base, "emplacement"]),
                    ("bottle_refs", [*base, "reference_bouteille"]),
                    ("models", [*raw_base, "model_diffuseur"]),
                    ("diffuser_refs", [*raw_base, "reference_diffuseur"]),
                    ("parfums", [*raw_base, "nom_parfum"]),
                    ("emplacements", [*raw_base, "emplacement"]),
                    ("bottle_refs", [*raw_base, "reference_bouteille"]),
                ]
            )

    recharges = payload.get("recharge_bouteille_effectuee")
    if isinstance(recharges, list):
        for index, _item in enumerate(recharges):
            base = ["recharge_bouteille_effectuee", index]
            raw_base = ["raw_payload", "recharge_bouteille_effectuee", index]
            paths.extend(
                [
                    ("parfums", [*base, "parfum"]),
                    ("emplacements", [*base, "emplacement"]),
                    ("bottle_refs", [*base, "reference_bouteille"]),
                    ("parfums", [*raw_base, "parfum"]),
                    ("emplacements", [*raw_base, "emplacement"]),
                    ("bottle_refs", [*raw_base, "reference_bouteille"]),
                ]
            )
    return paths


def sync_file(file_path: Path, references: dict[str, list[str]]) -> dict[str, Any]:
    payload = json.loads(file_path.read_text(encoding="utf-8"))
    changed = False
    replacements: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []

    for reference_key, path in iter_candidate_paths(payload):
        candidates = references.get(reference_key) or []
        if not candidates:
            continue
        original_value = get_from_path(payload, path)
        if original_value in (None, ""):
            continue
        matched_value, score = choose_reference(str(original_value), candidates)
        if matched_value is None:
            unmatched.append(
                {
                    "path": ".".join(str(part) for part in path),
                    "value": original_value,
                    "best_score": round(score, 4),
                }
            )
            continue
        if str(original_value) == matched_value:
            continue
        if set_in_path(payload, path, matched_value):
            replacements.append(
                {
                    "path": ".".join(str(part) for part in path),
                    "from": original_value,
                    "to": matched_value,
                    "score": round(score, 4),
                }
            )
            changed = True

    if changed:
        file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {
        "file": str(file_path),
        "changed": changed,
        "replacement_count": len(replacements),
        "replacements": replacements,
        "unmatched": unmatched,
    }


def sync_directory(directory: Path, references: dict[str, list[str]]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    if not directory.exists():
        return results
    for file_path in sorted(directory.rglob("*.json")):
        results.append(sync_file(file_path, references))
    return results


def build_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    changed_files = [result for result in results if result["changed"]]
    unmatched_files = [result for result in results if result["unmatched"]]
    return {
        "scanned_files": len(results),
        "changed_files": len(changed_files),
        "total_replacements": sum(result["replacement_count"] for result in results),
        "files_with_unmatched_values": len(unmatched_files),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync maintenance JSON reference values with canonical names from the Aromair Excel file."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--dir", dest="directories", action="append", type=Path)
    parser.add_argument("--report", type=Path, default=Path("data/excel_reference_sync_report.json"))
    args = parser.parse_args()

    directories = tuple(args.directories) if args.directories else DEFAULT_TARGET_DIRS
    references = load_excel_references(args.excel)
    results: list[dict[str, Any]] = []
    for directory in directories:
        results.extend(sync_directory(directory, references))

    report_payload = {
        "excel": str(args.excel),
        "directories": [str(directory) for directory in directories],
        "summary": build_summary(results),
        "results": results,
    }
    args.report.write_text(json.dumps(report_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report_payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
