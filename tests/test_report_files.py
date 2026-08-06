from __future__ import annotations

from openpyxl import load_workbook

from app.reporting.report_files import generate_report_excel, generate_report_pdf, is_valid_report_filename

_KPIS = {
    "totals": {
        "interventions": 28,
        "clients": 7,
        "volume_livre": 2800,
        "median_interventions_per_client": 3,
    },
    "top_clients": [
        {"name": "Boutique LARA", "count": 6},
        {"name": "Boumiza Square", "count": 6},
    ],
    "source_breakdown": [
        {"label": "Controle", "pct": 71.0},
        {"label": "Recharge", "pct": 29.0},
    ],
}
_REPORT_META = {
    "period": "week",
    "label": "Semaine du 29 juin 26 au 5 juil. 26",
    "range_start": "2026-06-29",
    "range_end": "2026-07-05",
}


def test_generate_report_pdf_creates_a_real_pdf_file():
    path = generate_report_pdf(_KPIS, _REPORT_META)
    try:
        assert path.exists()
        assert path.suffix == ".pdf"
        assert path.read_bytes().startswith(b"%PDF")
    finally:
        path.unlink(missing_ok=True)


def test_generate_report_excel_creates_a_workbook_with_expected_sheets():
    path = generate_report_excel(_KPIS, _REPORT_META)
    try:
        assert path.exists()
        assert path.suffix == ".xlsx"
        workbook = load_workbook(path)
        assert workbook.sheetnames == ["Resume", "Top clients", "Repartition"]
        assert workbook["Resume"]["A1"].value == "Rapport Auralys - Semaine du 29 juin 26 au 5 juil. 26"
        assert workbook["Top clients"]["A2"].value == "Boutique LARA"
        assert workbook["Repartition"]["B2"].value == 71.0
    finally:
        path.unlink(missing_ok=True)


def test_generated_filenames_are_always_valid():
    pdf_path = generate_report_pdf(_KPIS, _REPORT_META)
    excel_path = generate_report_excel(_KPIS, _REPORT_META)
    try:
        assert is_valid_report_filename(pdf_path.name)
        assert is_valid_report_filename(excel_path.name)
    finally:
        pdf_path.unlink(missing_ok=True)
        excel_path.unlink(missing_ok=True)


def test_is_valid_report_filename_rejects_path_traversal_and_junk():
    assert not is_valid_report_filename("../../etc/passwd")
    assert not is_valid_report_filename("..%2F..%2Fsecret.pdf")
    assert not is_valid_report_filename("")
    assert not is_valid_report_filename("rapport/with/slash.pdf")
