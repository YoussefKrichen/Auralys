from __future__ import annotations

import re
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Bind-mounted in prod (docker-compose.prod.yml: ./out:/app/out), so files
# written here survive a container restart instead of vanishing with the
# image layer -- same reasoning as the rest of the ./out convention.
REPORTS_DIR = Path("out/reports")

_FILENAME_RE = re.compile(r"^[A-Za-z0-9_.-]+$")


def is_valid_report_filename(filename: str) -> bool:
    """Guards the download route: only a name this module could have generated."""
    return bool(filename) and ".." not in filename and bool(_FILENAME_RE.fullmatch(filename))


def _build_filename(report_meta: dict[str, Any], extension: str) -> str:
    token = uuid4().hex[:8]
    return f"rapport_{report_meta['period']}_{report_meta['range_start']}_{token}.{extension}"


def generate_report_pdf(kpis: dict[str, Any], report_meta: dict[str, Any]) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORTS_DIR / _build_filename(report_meta, "pdf")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleAuralys",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#16324F"),
        spaceAfter=10,
        alignment=TA_LEFT,
    )
    h2_style = ParagraphStyle(
        "HeadingAuralys",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#1F4E79"),
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "BodyAuralys", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13
    )

    totals = kpis["totals"]
    story: list[Any] = [
        Paragraph(f"Rapport Auralys - {report_meta['label']}", title_style),
        Paragraph(f"Periode : {report_meta['range_start']} au {report_meta['range_end']}", body_style),
        Spacer(1, 0.4 * cm),
        Paragraph("Vue d'ensemble", h2_style),
        Paragraph(f"Interventions : {totals['interventions']}", body_style),
        Paragraph(f"Clients actifs : {totals['clients']}", body_style),
        Paragraph(f"Volume livre : {totals['volume_livre']} unites", body_style),
        Paragraph(f"Mediane interventions par client : {totals['median_interventions_per_client']}", body_style),
    ]

    if kpis.get("top_clients"):
        story.append(Paragraph("Top clients", h2_style))
        rows = [["Client", "Interventions"]] + [[row["name"], str(row["count"])] for row in kpis["top_clients"]]
        table = Table(rows, colWidths=[10 * cm, 4 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B8F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8D3CE")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F4F0")]),
                ]
            )
        )
        story.append(table)

    if kpis.get("source_breakdown"):
        story.append(Paragraph("Repartition par type d'intervention", h2_style))
        for row in kpis["source_breakdown"]:
            story.append(Paragraph(f"- {row['label']} : {row['pct']} %", body_style))

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        title=f"Rapport Auralys - {report_meta['label']}",
    )
    doc.build(story)
    return output_path


def generate_report_excel(kpis: dict[str, Any], report_meta: dict[str, Any]) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORTS_DIR / _build_filename(report_meta, "xlsx")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resume"
    bold = Font(bold=True)

    summary.append([f"Rapport Auralys - {report_meta['label']}"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([f"Periode : {report_meta['range_start']} au {report_meta['range_end']}"])
    summary.append([])

    totals = kpis["totals"]
    for label, value in (
        ("Interventions", totals["interventions"]),
        ("Clients actifs", totals["clients"]),
        ("Volume livre", totals["volume_livre"]),
        ("Mediane interventions par client", totals["median_interventions_per_client"]),
    ):
        summary.append([label, value])
        summary.cell(row=summary.max_row, column=1).font = bold
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 16

    if kpis.get("top_clients"):
        sheet = workbook.create_sheet("Top clients")
        sheet.append(["Client", "Interventions"])
        for cell in sheet[1]:
            cell.font = bold
        for row in kpis["top_clients"]:
            sheet.append([row["name"], row["count"]])
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 16

    if kpis.get("source_breakdown"):
        sheet = workbook.create_sheet("Repartition")
        sheet.append(["Type", "Pourcentage"])
        for cell in sheet[1]:
            cell.font = bold
        for row in kpis["source_breakdown"]:
            sheet.append([row["label"], row["pct"]])
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 16

    workbook.save(output_path)
    return output_path
